"""
評価結果記録スプレッドシートのテンプレートを作成するスクリプト
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_feedback_template():
    """フィードバック記録用のExcelテンプレートを作成"""

    wb = openpyxl.Workbook()

    # シート1: 評価結果記録
    ws_main = wb.active
    ws_main.title = "評価結果記録"

    # ヘッダー行の設定
    headers = [
        "候補者ID",
        "面談日",
        "AI総合スコア",
        "AIリスクレベル",
        "コミュニケーション",
        "ストレス耐性",
        "信頼性",
        "チームワーク",
        "Red Flags",
        "人間評価（5段階）",
        "アサイン結果",
        "問題の種類",
        "フィードバック",
        "備考"
    ]

    # ヘッダースタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ヘッダーを設定
    for col_num, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 列幅の設定
    column_widths = {
        'A': 12,  # 候補者ID
        'B': 12,  # 面談日
        'C': 14,  # AI総合スコア
        'D': 14,  # AIリスクレベル
        'E': 14,  # コミュニケーション
        'F': 14,  # ストレス耐性
        'G': 12,  # 信頼性
        'H': 14,  # チームワーク
        'I': 30,  # Red Flags
        'J': 16,  # 人間評価
        'K': 14,  # アサイン結果
        'L': 18,  # 問題の種類
        'M': 40,  # フィードバック
        'N': 30   # 備考
    }

    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width

    # サンプルデータを追加（参考用）
    sample_data = [
        ["001", "2026-03-15", 85, "低", 80, 75, 90, 85, "なし", 4, "成功", "なし", "技術力・コミュニケーション共に良好", ""],
        ["002", "2026-03-16", 45, "中", 50, 40, 45, 50, "視線を避ける、回答に矛盾", 2, "問題発生", "対人トラブル", "初期段階で懸念通りトラブル発生", ""],
        ["003", "2026-03-17", 70, "低", 75, 65, 70, 68, "なし", 3, "成功", "なし", "問題なく業務遂行中", ""],
    ]

    for row_num, data in enumerate(sample_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws_main.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 条件付き書式（AI総合スコアの色分け）
            if col_num == 3:  # AI総合スコア列
                if isinstance(value, (int, float)):
                    if value >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value >= 60:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 行の高さ調整
    ws_main.row_dimensions[1].height = 30
    for row in range(2, 100):
        ws_main.row_dimensions[row].height = 20

    # シート2: 集計・分析
    ws_analysis = wb.create_sheet("集計・分析")

    # タイトル
    ws_analysis['A1'] = "AI面談動画解析 - 集計・分析シート"
    ws_analysis['A1'].font = Font(size=14, bold=True)
    ws_analysis.merge_cells('A1:D1')

    # 基本統計
    ws_analysis['A3'] = "基本統計"
    ws_analysis['A3'].font = Font(size=12, bold=True)

    stats_labels = [
        "総評価件数",
        "平均AI総合スコア",
        "アサイン成功率",
        "問題発生率",
        "偽陽性率（AI低評価→成功）",
        "偽陰性率（AI高評価→問題）"
    ]

    for row, label in enumerate(stats_labels, 4):
        ws_analysis[f'A{row}'] = label
        ws_analysis[f'B{row}'] = "=COUNTA(評価結果記録!A:A)-1" if row == 4 else ""
        ws_analysis[f'A{row}'].font = Font(bold=True)

    # 相関分析
    ws_analysis['A12'] = "相関分析"
    ws_analysis['A12'].font = Font(size=12, bold=True)
    ws_analysis['A13'] = "AI評価と人間評価の相関係数"
    ws_analysis['B13'] = "=CORREL(評価結果記録!C:C, 評価結果記録!J:J)"

    # 注釈
    ws_analysis['A16'] = "※ サンプルデータは参考用です。実際のデータに置き換えてご使用ください。"
    ws_analysis['A17'] = "※ 相関係数は0.7以上が目標です。"
    ws_analysis['A16'].font = Font(italic=True, color="666666")
    ws_analysis['A17'].font = Font(italic=True, color="666666")

    # シート3: 使い方ガイド
    ws_guide = wb.create_sheet("使い方ガイド")

    guide_content = [
        ("AI面談動画解析 - 評価結果記録の使い方", Font(size=14, bold=True)),
        ("", None),
        ("1. 評価結果の記録", Font(size=12, bold=True)),
        ("   - 「評価結果記録」シートに、面談実施後すぐにAIの評価結果を記録してください。", None),
        ("   - 候補者IDは匿名化されたIDを使用してください（例: 001, 002...）。", None),
        ("   - AI総合スコアとカテゴリ別スコアはシステムの出力結果をそのまま転記してください。", None),
        ("", None),
        ("2. 人間の評価の記録", Font(size=12, bold=True)),
        ("   - 人間評価（5段階）: 面接官の総合評価を5段階で記録してください。", None),
        ("     5: 非常に良い、4: 良い、3: 普通、2: やや不安、1: 不安", None),
        ("", None),
        ("3. アサイン結果の記録", Font(size=12, bold=True)),
        ("   - アサイン結果: 成功/問題発生/その他 から選択してください。", None),
        ("   - 問題の種類: スキル不足/対人トラブル/途中辞退/なし から選択してください。", None),
        ("   - フィードバック: 詳細な状況や気づきを自由記述してください。", None),
        ("", None),
        ("4. 定期的な分析", Font(size=12, bold=True)),
        ("   - 「集計・分析」シートで、四半期ごとに精度を確認してください。", None),
        ("   - 相関係数が0.7以上であれば、AIの評価精度は良好です。", None),
        ("   - 偽陽性・偽陰性が多い場合は、システムの改善を検討してください。", None),
        ("", None),
        ("5. データの保護", Font(size=12, bold=True)),
        ("   - 個人情報保護のため、候補者の実名は記載しないでください。", None),
        ("   - ファイルはパスワード保護することを推奨します。", None),
    ]

    for row, (text, font) in enumerate(guide_content, 1):
        cell = ws_guide[f'A{row}']
        cell.value = text
        if font:
            cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_guide.column_dimensions['A'].width = 100

    # 保存
    output_path = "evaluation_tracker_template.xlsx"
    wb.save(output_path)
    print(f"フィードバックテンプレートを作成しました: {output_path}")
    print(f"作成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"シート数: 3 (評価結果記録、集計・分析、使い方ガイド)")

    return output_path

if __name__ == "__main__":
    create_feedback_template()
